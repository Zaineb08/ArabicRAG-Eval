#!/usr/bin/env python3
# =============================================================================
# ArabicRAG-Eval — evaluate_v2.py
# Implements Abdellah's 5 strategies (refinement guide, March 2026):
#   Strategy 1 — Cross-Evaluation (never let a model judge its own outputs)
#   Strategy 2 — No-Context Baseline analysis (RAG benefit score)
#   Strategy 3 — Context dependency split (GENERAL vs CONTEXT-DEPENDENT scores)
# =============================================================================

import os, sys, argparse, logging, warnings
from pathlib import Path
import pandas as pd


def load_env_files() -> None:
    """Load key=value pairs from common .env locations if present."""
    candidates = [
        Path.cwd() / ".env",
        Path(__file__).resolve().parent / ".env",
        Path(__file__).resolve().parent.parent / ".env",
    ]

    for env_path in candidates:
        if not env_path.exists():
            continue
        for line in env_path.read_text(encoding="utf-8").splitlines():
            raw = line.strip()
            if not raw or raw.startswith("#") or "=" not in raw:
                continue
            key, value = raw.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key:
                os.environ.setdefault(key, value)


load_env_files()

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("arabicrag_eval_v2")
warnings.filterwarnings("ignore", category=FutureWarning)

try:
    from ragas import evaluate
    from ragas.metrics import faithfulness, answer_relevancy, context_precision
    from ragas.llms import LangchainLLMWrapper
    from datasets import Dataset
except ImportError as e:
    log.error("Missing: %s — run: pip install ragas datasets openai langchain-openai", str(e))
    sys.exit(1)

MODELS      = ["llama", "qwen", "allam"]
METRICS     = [faithfulness, answer_relevancy, context_precision]
METRIC_NAMES = ["faithfulness", "answer_relevancy", "context_precision"]

MODEL_DISPLAY = {
    "llama": "LLaMA 3.3 70B",
    "qwen":  "Qwen3 32B",
    "allam": "ALLaM 2 7B (Arabic)",
}

# ── Strategy 1: Cross-evaluation judge assignment ─────────────────────────────
# Rule: never use a model to judge its own outputs (ArabicNLP 2024 cross-eval principle)
# LLaMA answers → judged by GPT-4o (different family)
# Qwen answers  → judged by GPT-4o (different family)
# ALLaM answers → judged by GPT-4o (different family)
# Default fallback: GPT-4o for all (acceptable since none are GPT)
JUDGE_MODEL = {
    "llama": "gpt-4o",   # GPT judges LLaMA — no conflict
    "qwen":  "gpt-4o",   # GPT judges Qwen — no conflict
    "allam": "gpt-4o",   # GPT judges ALLaM — no conflict
}

# Note: if testing GPT-4o answers in future, switch its judge to claude-3-opus


def get_ragas_llm(judge_model: str):
    """Return a RAGAS-compatible LLM wrapper for the given judge model."""
    try:
        from langchain_openai import ChatOpenAI
        llm = ChatOpenAI(model=judge_model, temperature=0)
        return LangchainLLMWrapper(llm)
    except Exception as e:
        log.warning("Could not load judge %s: %s. Falling back to default.", judge_model, str(e))
        return None


def build_ragas_dataset(df: pd.DataFrame, answer_col: str) -> Dataset:
    """Build HuggingFace Dataset for RAGAS from a results DataFrame."""
    records = []
    for _, row in df.iterrows():
        answer = str(row.get(answer_col, "")).strip()
        if answer.startswith("ERROR:") or not answer:
            continue
        records.append({
            "question":     str(row["question"]).strip(),
            "answer":       answer,
            "contexts":     [str(row["passage"]).strip()],
            "ground_truth": str(row["ground_truth_answer"]).strip(),
        })
    if not records:
        raise ValueError(f"No valid records for column '{answer_col}'")
    return Dataset.from_list(records)


def score_model(df: pd.DataFrame, model_key: str, use_no_context=False) -> pd.DataFrame:
    """Run RAGAS evaluation for one model, one condition (with/without context)."""
    suffix     = "_answer_no_context" if use_no_context else "_answer"
    answer_col = f"{model_key}{suffix}"
    judge      = JUDGE_MODEL.get(model_key, "gpt-4o")

    condition = "NO-CONTEXT baseline" if use_no_context else "WITH-CONTEXT"
    log.info("Scoring %s — %s (judge: %s)", MODEL_DISPLAY[model_key], condition, judge)

    ragas_ds = build_ragas_dataset(df, answer_col)
    log.info("  Records: %d", len(ragas_ds))

    # Strategy 1: set judge LLM explicitly
    ragas_llm = get_ragas_llm(judge)
    eval_kwargs = {"dataset": ragas_ds, "metrics": METRICS, "raise_exceptions": False}
    if ragas_llm:
        eval_kwargs["llm"] = ragas_llm

    result    = evaluate(**eval_kwargs)
    scores_df = result.to_pandas()

    valid_ids = df[
        ~df[answer_col].str.startswith("ERROR:", na=True) &
        df[answer_col].str.strip().ne("")
    ]["id"].values
    scores_df.insert(0, "id", valid_ids[:len(scores_df)])
    scores_df["model"]     = model_key
    scores_df["condition"] = "no_context" if use_no_context else "with_context"

    return scores_df


def compute_rag_benefit(with_scores: pd.DataFrame, no_scores: pd.DataFrame) -> pd.DataFrame:
    """
    Strategy 2: RAG benefit score = score WITH context - score WITHOUT context.
    A positive RAG benefit means the model genuinely uses the passage.
    A near-zero benefit means the model is relying on memory.
    """
    benefit_rows = []
    for metric in METRIC_NAMES:
        if metric in with_scores.columns and metric in no_scores.columns:
            benefit = with_scores[metric].mean() - no_scores[metric].mean()
            benefit_rows.append({
                "metric": metric,
                "with_context_mean": round(with_scores[metric].mean(), 4),
                "no_context_mean":   round(no_scores[metric].mean(), 4),
                "rag_benefit":       round(benefit, 4),
                "interpretation": (
                    "✅ Model uses context" if benefit > 0.05
                    else "⚠️ Low RAG benefit — possible memory reliance" if benefit >= 0
                    else "❌ Context hurts — possible prompt confusion"
                )
            })
    return pd.DataFrame(benefit_rows)


def split_by_context_dependency(df: pd.DataFrame, scores_df: pd.DataFrame) -> dict:
    """
    Strategy 3: Split scores by GENERAL vs CONTEXT-DEPENDENT questions.
    Returns dict with two DataFrames.
    """
    if "context_dependency" not in df.columns:
        log.warning("No context_dependency column found — skipping Strategy 3 split")
        return {}

    # Merge dependency tag into scores
    id_to_dep = dict(zip(df["id"].astype(str), df["context_dependency"]))
    scores_df["context_dependency"] = scores_df["id"].astype(str).map(id_to_dep)

    return {
        "GENERAL":           scores_df[scores_df["context_dependency"] == "GENERAL"],
        "CONTEXT-DEPENDENT": scores_df[scores_df["context_dependency"] == "CONTEXT-DEPENDENT"],
    }


def build_summary(all_with: dict, all_no: dict, all_benefits: dict) -> pd.DataFrame:
    """Build the final summary table with all conditions."""
    rows = []
    for model_key in all_with:
        with_df = all_with[model_key]
        no_df   = all_no.get(model_key, pd.DataFrame())
        row = {"Model": MODEL_DISPLAY[model_key]}
        for metric in METRIC_NAMES:
            if metric in with_df.columns:
                m_w = with_df[metric].mean()
                s_w = with_df[metric].std()
                row[f"{metric}_with"]    = round(m_w, 4)
                row[f"{metric}_with_disp"] = f"{m_w:.3f} ± {s_w:.3f}"
            if metric in no_df.columns:
                m_n = no_df[metric].mean()
                row[f"{metric}_no"]      = round(m_n, 4)
                row[f"{metric}_benefit"] = round(
                    row.get(f"{metric}_with", 0) - m_n, 4)
        rows.append(row)
    return pd.DataFrame(rows)


def export_excel(all_with, all_no, benefits, summary, dep_splits, output_path):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not available — skipping Excel export")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    thin      = Side(style="thin", color="CCCCCC")
    tborder   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gold_fill = PatternFill("solid", fgColor="FFF3CD")

    def style_header(ws, row_idx, n_cols):
        for c in range(1, n_cols+1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = tborder

    def style_data(ws, row_idx, n_cols, alt=False):
        for c in range(1, n_cols+1):
            cell = ws.cell(row=row_idx, column=c)
            if alt: cell.fill = PatternFill("solid", fgColor="EEF5FB")
            cell.border = tborder; cell.alignment = center

    def add_color_scale(ws, col_letter, n_rows):
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n_rows+1}",
            ColorScaleRule(
                start_type="min", start_color="FDECEA",
                mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                end_type="max", end_color="E8F8EF"))

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    headers = ["Model",
               "Faithfulness (W)", "Faithfulness (No-ctx)", "RAG Benefit (F)",
               "Ans. Relevancy (W)", "Ans. Relevancy (No-ctx)", "RAG Benefit (AR)",
               "Ctx. Precision (W)", "Ctx. Precision (No-ctx)", "RAG Benefit (CP)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for i, (_, row) in enumerate(summary.iterrows()):
        vals = [row.get("Model",""),
                row.get("faithfulness_with",""), row.get("faithfulness_no",""), row.get("faithfulness_benefit",""),
                row.get("answer_relevancy_with",""), row.get("answer_relevancy_no",""), row.get("answer_relevancy_benefit",""),
                row.get("context_precision_with",""), row.get("context_precision_no",""), row.get("context_precision_benefit","")]
        ws.append(vals)
        style_data(ws, i+2, len(headers), alt=(i%2==0))
    ws.column_dimensions["A"].width = 18
    for col in ["B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[col].width = 20

    # ── Sheet 2: RAG Benefit Analysis ─────────────────────────────────────────
    ws2 = wb.create_sheet("RAG Benefit (Strategy 2)")
    for model_key, benefit_df in benefits.items():
        ws2.append([f"── {MODEL_DISPLAY[model_key]} ──"])
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=12)
        ws2.append(["Metric","With Context","No Context","RAG Benefit","Interpretation"])
        style_header(ws2, ws2.max_row, 5)
        for _, r in benefit_df.iterrows():
            ws2.append([r["metric"], r["with_context_mean"], r["no_context_mean"],
                        r["rag_benefit"], r["interpretation"]])
        ws2.append([""])

    # ── Sheet 3: Context Dependency Split ─────────────────────────────────────
    ws3 = wb.create_sheet("Context Dependency (Strategy 3)")
    for model_key in all_with:
        scores_df = all_with[model_key]
        split = dep_splits.get(model_key, {})
        ws3.append([f"── {MODEL_DISPLAY[model_key]} ──"])
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=12)
        ws3.append(["Category","N","Faithfulness","Ans. Relevancy","Ctx. Precision"])
        style_header(ws3, ws3.max_row, 5)
        for cat in ["GENERAL", "CONTEXT-DEPENDENT"]:
            sub = split.get(cat, pd.DataFrame())
            if len(sub) > 0:
                ws3.append([cat, len(sub),
                    round(sub["faithfulness"].mean(), 3) if "faithfulness" in sub.columns else "",
                    round(sub["answer_relevancy"].mean(), 3) if "answer_relevancy" in sub.columns else "",
                    round(sub["context_precision"].mean(), 3) if "context_precision" in sub.columns else ""])
        ws3.append([""])

    # ── Per-model WITH context sheets ─────────────────────────────────────────
    for model_key, scores_df in all_with.items():
        ws_m = wb.create_sheet(f"{MODEL_DISPLAY[model_key][:20]} (w/ctx)")
        cols = ["id","faithfulness","answer_relevancy","context_precision","context_dependency"]
        hdrs = ["ID","Faithfulness","Ans. Relevancy","Ctx. Precision","Dependency"]
        ws_m.append(hdrs)
        style_header(ws_m, 1, len(hdrs))
        for i, (_, row) in enumerate(scores_df.iterrows()):
            ws_m.append([row.get(c,"") for c in cols])
            style_data(ws_m, i+2, len(hdrs), alt=(i%2==0))
        for col_letter in ["B","C","D"]:
            add_color_scale(ws_m, col_letter, len(scores_df))

    wb.save(output_path)
    log.info("Excel report saved: %s", output_path)


def run_evaluation(results_path, output_prefix, models=None):
    log.info("Loading results: %s", results_path)
    df = pd.read_csv(results_path, encoding="utf-8")
    df["id"] = df["id"].astype(str)
    log.info("Rows: %d", len(df))

    required = {"id","passage","question","ground_truth_answer"}
    if not required.issubset(df.columns):
        log.error("Missing columns: %s", required - set(df.columns))
        sys.exit(1)

    if not os.environ.get("OPENAI_API_KEY"):
        log.warning("OPENAI_API_KEY not set — RAGAS will use default judge")

    active_models = models or [m for m in MODELS if f"{m}_answer" in df.columns]

    all_with     = {}
    all_no       = {}
    all_benefits = {}
    dep_splits   = {}

    for model_key in active_models:
        # WITH context scores
        try:
            with_scores = score_model(df, model_key, use_no_context=False)
            all_with[model_key] = with_scores
        except Exception as e:
            log.error("WITH-context scoring failed for %s: %s", model_key, e)
            continue

        # NO context scores (Strategy 2)
        no_col = f"{model_key}_answer_no_context"
        if no_col in df.columns:
            try:
                no_scores = score_model(df, model_key, use_no_context=True)
                all_no[model_key] = no_scores
                # RAG benefit
                all_benefits[model_key] = compute_rag_benefit(with_scores, no_scores)
            except Exception as e:
                log.warning("NO-context scoring failed for %s: %s", model_key, e)

        # Context dependency split (Strategy 3)
        split = split_by_context_dependency(df, with_scores.copy())
        dep_splits[model_key] = split

        # Log quick results
        for metric in METRIC_NAMES:
            if metric in with_scores.columns:
                w = with_scores[metric].mean()
                n = all_no.get(model_key, pd.DataFrame()).get(metric, pd.Series([float("nan")])).mean()
                log.info("  %-6s %-22s  WITH=%.3f  NO-CTX=%.3f  BENEFIT=%.3f",
                         model_key.upper(), metric, w, n, w-n)

    # Save per-triplet CSV
    per_triplet = pd.concat(
        [s.assign(condition="with_context") for m,s in all_with.items()] +
        [s.assign(condition="no_context")   for m,s in all_no.items()],
        ignore_index=True)
    per_triplet.to_csv(f"{output_prefix}_PerTriplet.csv", index=False, encoding="utf-8")

    # Save RAG benefit CSV
    if all_benefits:
        benefit_rows = []
        for model_key, bdf in all_benefits.items():
            bdf["model"] = MODEL_DISPLAY[model_key]
            benefit_rows.append(bdf)
        pd.concat(benefit_rows).to_csv(f"{output_prefix}_RAGBenefit.csv", index=False, encoding="utf-8")

    # Summary
    summary = build_summary(all_with, all_no, all_benefits)
    summary.to_csv(f"{output_prefix}_Summary.csv", index=False, encoding="utf-8")

    # Console summary
    print("\n" + "="*72)
    print("  ArabicRAG-Eval v2 — Results Summary")
    print("="*72)
    print(f"  {'Model':<22} {'Faith (W)':>10} {'Faith (No)':>11} {'Benefit':>9}")
    print("  " + "-"*60)
    for _, row in summary.iterrows():
        print(f"  {row['Model']:<22} "
              f"{row.get('faithfulness_with',float('nan')):>10.3f} "
              f"{row.get('faithfulness_no',float('nan')):>11.3f} "
              f"{row.get('faithfulness_benefit',float('nan')):>9.3f}")
    print("="*72)
    print()
    print("  Strategy 3 — Context Dependency Split:")
    for model_key in active_models:
        split = dep_splits.get(model_key, {})
        for cat in ["GENERAL","CONTEXT-DEPENDENT"]:
            sub = split.get(cat, pd.DataFrame())
            if len(sub) > 0 and "faithfulness" in sub.columns:
                print(f"  {MODEL_DISPLAY[model_key]:<22} {cat:<22} faithfulness={sub['faithfulness'].mean():.3f}")
    print("="*72 + "\n")

    # Excel
    export_excel(all_with, all_no, all_benefits, summary, dep_splits,
                 f"{output_prefix}_Report.xlsx")

    log.info("All outputs saved with prefix: %s", output_prefix)


def parse_args():
    parser = argparse.ArgumentParser(
        description="ArabicRAG-Eval v2 evaluation with cross-judge, baseline, and dependency split",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--results", default="ArabicRAG-Eval_Results_v2.csv")
    parser.add_argument("--output",  default="ArabicRAG-Eval_Scores_v2")
    parser.add_argument("--models",  nargs="+", choices=MODELS, default=None)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run_evaluation(args.results, args.output, args.models)
